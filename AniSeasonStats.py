import requests
import pandas as pd
import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

API_URL = "https://graphql.anilist.co"

def fetch_anime_data(username):
    query = """
    query ($username: String) {
    MediaListCollection(userName: $username, type: ANIME) {
        lists {
        entries {
            media {
            title {
                romaji
            }
            season
            seasonYear
            averageScore
            popularity
            }
            score
            status
            }
        }
    }
    
    }
    """
    variables = {"username": username}
    response = requests.post(API_URL, json={"query": query, "variables": variables})
    if response.status_code == 200:
        return response.json()
    else:
        print(f"Error: {response.status_code}")
        print(response.json())
        return None

def calculate_weighted_mean(scores, popularities, season_mean, global_mean, seen_count, total_count, wp=0.6, wa=0.4):
    if any(pop for score, pop in zip(scores, popularities) if score > 0):
        popularity_weighted_mean = (
            sum(score * pop for score, pop in zip(scores, popularities) if score > 0) /
            sum(pop for score, pop in zip(scores, popularities) if score > 0)
        )
    else:
        popularity_weighted_mean = 0

    activity_ratio = seen_count / total_count if total_count > 0 else 0
    activity_weighted_mean = (
        activity_ratio * season_mean + (1 - activity_ratio) * global_mean
    )

    weighted_mean = round(
        wp * popularity_weighted_mean + wa * activity_weighted_mean, 2
    )

    return weighted_mean

def season_order(season):
    order = {"WINTER": 1, "SPRING": 2, "SUMMER": 3, "FALL": 4}
    return order.get(season, 5)  # Default to a high number for unknown seasons

def calculate_statistics(data, global_mean=7.0, total_count=50):
    if not data or "data" not in data or "MediaListCollection" not in data["data"]:
        return []

    stats_by_season = {}
    for list_data in data["data"]["MediaListCollection"]["lists"]:
        for entry in list_data["entries"]:
            if entry.get("status") == "COMPLETED":
                media = entry["media"]
                season = media.get("season")
                year = media.get("seasonYear")
                if not season or not year or year < 2006:
                    continue

                key = (year, season)
                if key not in stats_by_season:
                    stats_by_season[key] = {"scores": [], "popularities": [], "titles": set()}

                score = entry.get("score", 0)
                title = media["title"]["romaji"]
                stats_by_season[key]["titles"].add((score, title))
                if score:
                    stats_by_season[key]["scores"].append(score)
                popularity = media.get("popularity", 0)
                stats_by_season[key]["popularities"].append(popularity)

    stats = []
    for (year, season), details in sorted(stats_by_season.items(), key=lambda x: (x[0][0], season_order(x[0][1]))):
        sorted_titles = sorted(details["titles"], key=lambda x: x[1])
        formatted_titles = [f"{score} - {title}" for score, title in sorted_titles]

        scores = details["scores"]
        popularities = details["popularities"]
        mean_score = round(sum(scores) / len(scores), 2) if scores else 0

        weighted_mean = calculate_weighted_mean(
            scores, popularities, season_mean=mean_score, global_mean=global_mean,
            seen_count=len(details["titles"]), total_count=total_count
        )

        stats.append({
            "season": season,
            "year": year,
            "anime_count": len(details["titles"]),
            "mean_score": mean_score,
            "weighted_mean": weighted_mean,
            "anime_list": formatted_titles,
        })
    return stats

def save_to_excel_with_formatting(stats, filename):
    rows = []
    for stat in stats:
        row = {
            "Season": stat["season"],
            "Year": stat["year"],
            "Anime Count": stat["anime_count"],
            "Anime Score": stat["mean_score"],
            "Weighted Score": stat["weighted_mean"],
        }
        for i, anime in enumerate(stat["anime_list"]):
            row[f"Anime {i + 1}"] = anime
        rows.append(row)

    df = pd.DataFrame(rows)
    current_date = datetime.datetime.now().strftime("%Y-%m-%d")
    filepath = f"{current_date}_{filename}.xlsx"
    df.to_excel(filepath, index=False)

    wb = load_workbook(filepath)
    ws = wb.active

    apply_custom_formatting(ws, "C", top=2, bottom=2)
    apply_custom_formatting(ws, "D", top=2, bottom=2)
    apply_custom_formatting(ws, "E", top=2, bottom=2)
    apply_autofit(ws, ["C", "D", "E"])

    wb.save(filepath)
    print(f"Data saved to {filepath}")

def apply_custom_formatting(sheet, column_letter, top=2, bottom=2, ignore_zeros=True):
    values = [
        (row, sheet[f"{column_letter}{row}"].value)
        for row in range(2, sheet.max_row + 1)
        if not ignore_zeros or sheet[f"{column_letter}{row}"].value != 0
    ]

    sorted_values = sorted(values, key=lambda x: x[1])
    lowest = sorted_values[:bottom]
    highest = sorted_values[-top:]

    green_fill = PatternFill(start_color="32CD32", end_color="32CD32", fill_type="solid")
    green_font = Font(color="0F703B")
    red_fill = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")
    red_font = Font(color="9C0006")

    for row, _ in lowest:
        cell = sheet[f"{column_letter}{row}"]
        cell.fill = red_fill
        cell.font = red_font
    for row, _ in highest:
        cell = sheet[f"{column_letter}{row}"]
        cell.fill = green_fill
        cell.font = green_font

def apply_autofit(sheet, columns):
    for column in columns:
        max_length = 0
        col_index = sheet[column + "1"].column
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=col_index, max_col=col_index):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[get_column_letter(col_index)].width = max_length + 2

if __name__ == "__main__":
    username = input("Please enter your AniList username: ").strip()
    if not username:
        print("No username entered. Exiting program.")
    else:
        data = fetch_anime_data(username)
        stats = calculate_statistics(data)
        save_to_excel_with_formatting(stats, "anime_season_stats")
